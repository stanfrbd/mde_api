import json
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils.api_request import get_token, read_secrets

# disable ssl warning in case of proxy like Zscaler which breaks ssl...
requests.packages.urllib3.disable_warnings()

# Current date
now = datetime.now()
today = now.strftime("%Y-%m-%d-%H_%M_%S")

def export_to_csv(data):
    filename = today + "-mde-api-results.csv"
    with open(filename, "a") as f:
        f.write("machine,os,sensor_status,onboarding_status,verification")
        for row in data:
            f.write(",".join(row) + "\n")

def export_to_excel(data):
    filename = today + "-mde-api-results.xlsx"
    wb = Workbook()
    ws = wb.active

    # Set file headers
    ws.append(["machine", "os", "sensor_status", "onboarding_status", "verification"])

    # Define conditional formatting colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2):
        # Sensor column (column C)
        if row[2].value.startswith("Inactive"):
            row[2].fill = orange_fill
        elif row[2].value.startswith("Active"):
            row[2].fill = green_fill

        # Onboard column (column D)
        if row[3].value.startswith("CanBeOnboarded"):
            row[3].fill = orange_fill
        elif row[3].value.startswith("Onboarded"):
            row[3].fill = green_fill
        elif row[3].value.startswith("InsufficientInfo"):
            row[3].fill = light_orange_fill
        elif row[3].value.startswith("Unsupported"):
            row[3].fill = red_fill
        
        # Verification (column E)
        if row[4].value.startswith("NotFullyOnboarded"):
            row[4].fill = orange_fill
        elif row[4].value.startswith("FullyOnboarded"):
            row[4].fill = green_fill
        elif row[4].value.startswith("InsufficientInfo"):
            row[4].fill = light_orange_fill
        elif row[4].value.startswith("Unsupported"):
            row[4].fill = red_fill

    # Add table with filters
    table = Table(displayName="ResultsTable", ref="A1:E{}".format(ws.max_row))
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Adjust columns width
    for column in ws.iter_cols():
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    print("Successfully created {}".format(filename))

# List Windows Servers onboarding status (sensor and onboarding)
def get_windows_servers_onboarding_status():

    secrets = read_secrets()
    token = get_token(secrets)

    headers = { 
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'Authorization': "Bearer {}".format(token)
    }

    proxies = { 'http': secrets['proxyUrl'], 'https': secrets['proxyUrl'] }

    # url = "https://api.securitycenter.windows.com/api/machines"
    # filter for Windows Servers only
    url = "https://api.securitycenter.windows.com/api/machines?$filter=startswith(osPlatform,'WindowsServer')"
    response = requests.get(url, headers=headers, proxies=proxies, verify=False)
    json_response = json.loads(response.content)
    with open("machines.json", "w") as f:
        f.write(response.text)

    # f = open("machines.json", "r")
    # json_response = json.loads(f.read())
    # f.close()

    api_results = []
    cpt_api_results = 0
    cpt_active_onboarded = 0
    cpt_canbeonboarded = 0
    cpt_inactive = 0
    cpt_unsupported = 0
    cpt_insufficientinfo = 0
    for machine in json_response["value"]:
        #if machine["osPlatform"].__contains__("WindowsServer"):
        # print("{},{},{},{}".format(machine["computerDnsName"], machine["osPlatform"], machine["healthStatus"], machine["onboardingStatus"]))

        if machine["healthStatus"] == "Active" and machine["onboardingStatus"] == "Onboarded":
            verification = "FullyOnboarded"
            cpt_active_onboarded += 1

        elif machine["onboardingStatus"] == "CanBeOnboarded":
            verification = "NotFullyOnboarded"
            cpt_canbeonboarded += 1
        
        elif machine["onboardingStatus"] == "Unsupported":
            verification = "Unsupported"
            cpt_unsupported += 1
        
        elif machine["onboardingStatus"] == "InsufficientInfo":
            verification = "InsufficientInfo"
            cpt_insufficientinfo += 1
        
        elif machine["healthStatus"] == "Inactive":
            verification = "NotFullyOnboarded"
            cpt_inactive += 1
        
        else:
           verification = "NotFullyOnboarded" 

        cpt_api_results += 1

        api_results.append([machine["computerDnsName"], machine["osPlatform"], machine["healthStatus"], machine["onboardingStatus"], verification])

    print("{} fully onboarded, {} can be onboarded, {} unsupported, {} no info, {} inactive.".format(cpt_active_onboarded, cpt_canbeonboarded, cpt_unsupported, cpt_insufficientinfo, cpt_inactive))
    print("Processed {} API results.".format(cpt_api_results))
    export_to_excel(api_results)