#!/usr/bin/env python3

# Stanislas M. 2023-12-26

"""
usage: mde_api.py [-h] {token,machines,indicators,vulnerabilities,users,software} ...

List machines, offboard or token

positional arguments:
  {token,machines,indicators,vulnerabilities,users,software}
    token               Get token
    machines            Perform actions on machines
    indicators          Perform actions on indicators
    vulnerabilities     Perform actions on vulnerabilities
    users               Perform actions on users
    software            Perform actions on software

options:
  -h, --help            show this help message and exit
"""

import json
import requests
import argparse
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# disable ssl warning in case of proxy like Zscaler which breaks ssl...
requests.packages.urllib3.disable_warnings()

# Current date
now = datetime.now()
today = now.strftime("%Y-%m-%d-%H_%M_%S")

def export_to_csv(data):
    filename = today + "-mde-api-results.csv"
    with open(filename, "a") as f:
        f.write("machine,os,sensor_status,onboarding_status,verification")
        for row in data:
            f.write(",".join(row) + "\n")

def export_to_excel(data):
    filename = today + "-mde-api-results.xlsx"
    wb = Workbook()
    ws = wb.active

    # Set file headers
    ws.append(["machine", "os", "sensor_status", "onboarding_status", "verification"])

    # Define conditional formatting colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    # Add data to the worksheet
    for row in data:
        ws.append(row)

    # Apply conditional formatting
    for row in ws.iter_rows(min_row=2):
        # Sensor column (column C)
        if row[2].value.startswith("Inactive"):
            row[2].fill = orange_fill
        elif row[2].value.startswith("Active"):
            row[2].fill = green_fill

        # Onboard column (column D)
        if row[3].value.startswith("CanBeOnboarded"):
            row[3].fill = orange_fill
        elif row[3].value.startswith("Onboarded"):
            row[3].fill = green_fill
        elif row[3].value.startswith("InsufficientInfo"):
            row[3].fill = light_orange_fill
        elif row[3].value.startswith("Unsupported"):
            row[3].fill = red_fill
        
        # Verification (column E)
        if row[4].value.startswith("NotFullyOnboarded"):
            row[4].fill = orange_fill
        elif row[4].value.startswith("FullyOnboarded"):
            row[4].fill = green_fill
        elif row[4].value.startswith("InsufficientInfo"):
            row[4].fill = light_orange_fill
        elif row[4].value.startswith("Unsupported"):
            row[4].fill = red_fill

    # Add table with filters
    table = Table(displayName="ResultsTable", ref="A1:E{}".format(ws.max_row))
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Adjust columns width
    for column in ws.iter_cols():
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(filename)
    print("Successfully created {}".format(filename))

# Read secrets from "secrets.json"
def read_secrets():
    with open('secrets.json') as f:
        secrets = json.load(f)
    if not (key in secrets for key in ['tenantId', 'appId', 'appSecret']):
        raise ValueError('Error: Invalid secrets file')
    return secrets

# Make Post request to get AAD Token
def get_token(secrets):

    url = "https://login.microsoftonline.com/{}/oauth2/token".format(secrets['tenantId'])

    resourceAppIdUri = 'https://api.securitycenter.microsoft.com'

    body = {
        'resource': resourceAppIdUri,
        'client_id': secrets['appId'],
        'client_secret': secrets['appSecret'],
        'grant_type': 'client_credentials'
    }

    proxies = { 'http': secrets['proxyUrl'], 'https': secrets['proxyUrl'] }

    try:
        response = requests.post(url, data=body, proxies=proxies, verify=False)
        json_response = json.loads(response.content)
    except Exception as err:
        print("Error: " + str(err))
    try:
        aad_token = json_response["access_token"]
    except:
        print("Error: Unable to retrieve token")
        aad_token = "invalid"
    return aad_token

def list_alerts():
    # Code to list alerts
    pass

# Offboard a machine using its ID
def offboard_machine(machine_id):
    secrets = read_secrets()
    token = get_token(secrets)

    headers = { 
        'Content-Type': 'application/json',
        'Accept':'application/json',
        'Authorization': "Bearer {}".format(token)
    }

    body = {
        'Comment': 'Offboard machine by automation',
    }

    proxies = { 'http': secrets['proxyUrl'], 'https': secrets['proxyUrl'] }

    url = "https://api.securitycenter.microsoft.com/api/machines/{}/offboard".format(machine_id)

    try:
        response = requests.post(url, json=body, headers=headers, proxies=proxies, verify=False)
        json_response = json.loads(response.content)
    
    except Exception as err:
        print("Unable to proceed to offboarding request: {}".format(str(err)))

    with open("offboard.json", "w") as f:
        f.write(response.text)
    
    print(json.dumps(json_response, indent=4))

# List Windows Servers onboarding status (sensor and onboarding)
def get_server_onboarding_status():

    secrets = read_secrets()
    token = get_token(secrets)

    headers = { 
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'Authorization': "Bearer {}".format(token)
    }

    proxies = { 'http': secrets['proxyUrl'], 'https': secrets['proxyUrl'] }

    # url = "https://api.securitycenter.windows.com/api/machines"
    # filter for Windows Servers only
    url = "https://api.securitycenter.windows.com/api/machines?$filter=startswith(osPlatform,'WindowsServer')"
    response = requests.get(url, headers=headers, proxies=proxies, verify=False)
    json_response = json.loads(response.content)
    with open("machines.json", "w") as f:
        f.write(response.text)

    # f = open("machines.json", "r")
    # json_response = json.loads(f.read())
    # f.close()

    api_results = []
    cpt_api_results = 0
    cpt_active_onboarded = 0
    cpt_canbeonboarded = 0
    cpt_inactive = 0
    cpt_unsupported = 0
    cpt_insufficientinfo = 0
    for machine in json_response["value"]:
        #if machine["osPlatform"].__contains__("WindowsServer"):
        # print("{},{},{},{}".format(machine["computerDnsName"], machine["osPlatform"], machine["healthStatus"], machine["onboardingStatus"]))

        if machine["healthStatus"] == "Active" and machine["onboardingStatus"] == "Onboarded":
            verification = "FullyOnboarded"
            cpt_active_onboarded += 1

        elif machine["onboardingStatus"] == "CanBeOnboarded":
            verification = "NotFullyOnboarded"
            cpt_canbeonboarded += 1
        
        elif machine["onboardingStatus"] == "Unsupported":
            verification = "Unsupported"
            cpt_unsupported += 1
        
        elif machine["onboardingStatus"] == "InsufficientInfo":
            verification = "InsufficientInfo"
            cpt_insufficientinfo += 1
        
        elif machine["healthStatus"] == "Inactive":
            verification = "NotFullyOnboarded"
            cpt_inactive += 1
        
        else:
           verification = "NotFullyOnboarded" 

        cpt_api_results += 1

        api_results.append([machine["computerDnsName"], machine["osPlatform"], machine["healthStatus"], machine["onboardingStatus"], verification])

    print("{} fully onboarded, {} can be onboarded, {} unsupported, {} no info, {} inactive.".format(cpt_active_onboarded, cpt_canbeonboarded, cpt_unsupported, cpt_insufficientinfo, cpt_inactive))
    print("Processed {} API results.".format(cpt_api_results))
    export_to_excel(api_results)

def show_token():
    print("\nHere is your JWT token, test it here: https://jwt.ms/\n")
    print(get_token(read_secrets()))
    print("")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='List machines, offboard or token')
    
    subparsers = parser.add_subparsers(dest='command')

    # Create the parser for the "token" command
    token_parse = subparsers.add_parser('token', help='Get token')

    # create the parser for the "machines" command
    machines_parser = subparsers.add_parser('machines', help='Perform actions on machines')
    machines_subparsers = machines_parser.add_subparsers(dest='subcommand')

    # create the argument for the "report" command
    machines_subparsers.add_parser('report', help='Generate Excel report')

    # create the parser for the "offboard" command
    offboard_parser = machines_subparsers.add_parser('offboard', help='Offboard machine')
    offboard_parser.add_argument('machineid', help='Machine ID')

    # create the parser for the "isolate" command
    isolate_parser = machines_subparsers.add_parser('isolate', help='Isolate machine')
    isolate_parser.add_argument('machineid', help='Machine ID')

    # create the parser for the "indicators" command
    indicators_parser = subparsers.add_parser('indicators', help='Perform actions on indicators')
    indicators_subparsers = indicators_parser.add_subparsers(dest='subcommand')

    # create the parser for the "add" command
    add_parser = indicators_subparsers.add_parser('add', help='Add an indicator')
    add_parser.add_argument('indicator', help='Indicator')

    # create the parser for the "vulnerabilities" command
    vulnerabilities_parser = subparsers.add_parser('vulnerabilities', help='Perform actions on vulnerabilities')
    vulnerabilities_subparsers = vulnerabilities_parser.add_subparsers(dest='subcommand')

    # create the parser for the "get_devices" command
    get_devices_parser = vulnerabilities_subparsers.add_parser('get_devices', help='Get devices for a CVE')
    get_devices_parser.add_argument('cveid', help='CVE ID')

    # create the parser for the "users" command
    users_parser = subparsers.add_parser('users', help='Perform actions on users')
    users_subparsers = users_parser.add_subparsers(dest='subcommand')

    # create the parser for the "list" command
    list_parser = users_subparsers.add_parser('list', help='List all users')

    # create the parser for the "get" command
    get_parser = users_subparsers.add_parser('get', help='Get a user')
    get_parser.add_argument('userUpn', help='User UPN')

    # create the parser for the "software" command
    software_parser = subparsers.add_parser('software', help='Perform actions on software')
    software_subparsers = software_parser.add_subparsers(dest='subcommand')

    args = parser.parse_args()

    try: 
        if args.command == 'machines':
            if args.subcommand == 'list':
                print("Listing all machines...")
            elif args.subcommand == 'offboard':
                print(f"Offboarding machine with ID: {args.machineid}")
                offboard_machine(args.machineid)
            elif args.subcommand == 'report':
                get_server_onboarding_status()
            elif args.subcommand == 'isolate':
                print(f"Isolating machine with ID: {args.machineid}")
        
        elif args.command == 'token':
            show_token()
        
        
    except Exception as err:
        print("General error: {} check your secrets file and App Registration.".format(str(err))) 
        exit(1)
